from __future__ import annotations

from pathlib import Path
import re
import yaml
import hashlib
from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo
from typing import Any, Dict, Optional

import openpyxl


DOCS_DIR = Path("docs")
RAW_DIR = DOCS_DIR / "cases_raw"   # 投稿源文件（由 xlsx 自动生成）
OUT_DIR = DOCS_DIR / "cases"       # 自动生成的展示页面

INDEX_FILE = OUT_DIR / "index.md"
BY_UNI_FILE = OUT_DIR / "by-university.md"
BY_MAJOR_FILE = OUT_DIR / "by-major.md"

# 新增页面
EXPERIENCE_FILE = DOCS_DIR / "experience.md"

# “学长学姐说”目录（长文）
SENIORS_DIR = DOCS_DIR / "seniors"
SENIORS_INDEX_FILE = SENIORS_DIR / "index.md"

FRONT_MATTER_RE = re.compile(r"^---\s*\n(.*?)\n---\s*\n", re.DOTALL)
FIRST_H1_RE = re.compile(r"^\s*#\s+(.+?)\s*$", re.MULTILINE)


# =========================================================
# 1) XLSX -> cases_raw/*.md
# =========================================================

def xnorm(v: Any) -> str:
    """
    规范化 xlsx 单元格内容：
    - None / 空白 / (空) / （空） -> ""
    - 其余转为去空格字符串
    """
    if v is None:
        return ""

    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")

    s = str(v).strip()

    # 腾讯问卷“未填写”的常见表示
    EMPTY_MARKERS = {
        "",
        "(空)",
        "（空）",
        "-",
        "—",
        "无",
        "NULL",
        "null",
    }

    if s in EMPTY_MARKERS:
        return ""

    return s


def yaml_quote(s: str) -> str:
    """最小转义，确保 YAML/front matter 不炸。"""
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{s}"'


def normalize_track(raw: str) -> str:
    """
    归一选科：物理类/历史类 -> 物理/历史
    """
    t = xnorm(raw)
    if not t:
        return ""
    if "物理" in t:
        return "物理"
    if "历史" in t:
        return "历史"
    return t


def find_col(headers: list[str], patterns: list[str]) -> Optional[int]:
    for pat in patterns:
        reg = re.compile(pat)
        for i, h in enumerate(headers):
            if reg.search(h):
                return i
    return None


def build_column_map(headers: list[str]) -> Dict[str, int]:
    """
    适配【深中飞跃手册案例投稿】最新版问卷表头
    """

    col: Dict[str, Optional[int]] = {}

    col["nickname"] = find_col(headers, [r"昵称"])
    col["exam_year"] = find_col(headers, [r"高考年份"])
    col["track"] = find_col(headers, [r"选科", r"科目"])

    col["sz_mock1_rank"] = find_col(headers, [r"深一模"])
    col["sz_mock2_rank"] = find_col(headers, [r"深二模"])

    col["gaokao_score"] = find_col(headers, [r"高考分数", r"高考成绩"])
    col["gaokao_rank"] = find_col(headers, [r"高考.*排名", r"省排名", r"省.*位次"])

    col["university"] = find_col(headers, [r"录取院校", r"录取学校"])
    col["major"] = find_col(headers, [r"录取专业", r"就读专业"])

    col["university_review"] = find_col(headers, [r"院校评价"])
    col["major_review"] = find_col(headers, [r"专业评价"])
    col["advice"] = find_col(headers, [r"学弟学妹", r"建议"])

    col["submit_time"] = find_col(headers, [r"提交.*时间"])

    # 必填字段校验
    required = [
        "exam_year",
        "track",
        "gaokao_score",
        "gaokao_rank",
        "university",
        "major",
    ]
    missing = [k for k in required if col.get(k) is None]
    if missing:
        raise RuntimeError(
            "Excel 表头缺少必要字段（可能是问卷又改名了）：\n"
            + "\n".join(f"- {k}" for k in missing)
        )

    return {k: int(v) for k, v in col.items() if v is not None}


def row_to_meta(row: tuple, colmap: Dict[str, int]) -> Dict[str, str]:
    def get(key: str) -> str:
        idx = colmap.get(key)
        if idx is None or idx >= len(row):
            return ""
        return xnorm(row[idx])

    meta: Dict[str, str] = {
        "nickname": get("nickname"),
        "exam_year": get("exam_year"),
        "track": normalize_track(get("track")),
        "sz_mock1_rank": get("sz_mock1_rank"),
        "sz_mock2_rank": get("sz_mock2_rank"),
        "gaokao_score": get("gaokao_score"),
        "gaokao_rank": get("gaokao_rank"),
        "university": get("university"),
        "major": get("major"),
        "university_review": get("university_review"),
        "major_review": get("major_review"),
        "advice": get("advice"),
        "_submit_time": get("submit_time"),
    }
    return meta


def make_submission_filename(meta: Dict[str, str], idx: int) -> str:
    """
    生成稳定且不依赖昵称的文件名
    """
    ts = meta.get("_submit_time") or datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S")
    safe_ts = re.sub(r"[^0-9A-Za-z]+", "-", ts).strip("-")
    base = "|".join(
        [
            safe_ts,
            str(idx),
            meta.get("exam_year", ""),
            meta.get("gaokao_score", ""),
            meta.get("gaokao_rank", ""),
            meta.get("university", ""),
            meta.get("major", ""),
        ]
    )
    h = hashlib.sha1(base.encode("utf-8")).hexdigest()[:10]
    return f"submission-{safe_ts}-{idx:04d}-{h}.md"


def write_case_raw_md(meta: Dict[str, str], out_path: Path) -> None:
    """
    输出为 cases_raw 的 front matter。正文留空即可。
    选填为空写 ""，页面显示 NULL / Anonymous 由 build 逻辑控制。
    """
    keys = [
        "nickname",
        "exam_year",
        "track",
        "sz_mock1_rank",
        "sz_mock2_rank",
        "gaokao_score",
        "gaokao_rank",
        "university",
        "major",
        "university_review",
        "major_review",
        "advice",
    ]

    lines = ["---"]
    for k in keys:
        lines.append(f"{k}: {yaml_quote(meta.get(k, ''))}")
    lines.append("---")
    lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def import_xlsx_to_cases_raw() -> int:
    """
    约定：
    - docs/ 下必须且只能有一个 xlsx
    - 导入前清空 docs/cases_raw 下所有 .md（避免历史残留）
    返回：写入的 md 数量
    """
    xlsx_files = list(DOCS_DIR.glob("*.xlsx"))
    if len(xlsx_files) != 1:
        raise RuntimeError(f"docs/ 目录下应当且只能存在一个 xlsx 文件，当前找到 {len(xlsx_files)} 个。")

    xlsx_path = xlsx_files[0]

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    # 清空旧 md
    for p in RAW_DIR.glob("*.md"):
        p.unlink()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    it = ws.iter_rows(values_only=True)
    header_row = next(it, None)
    if not header_row:
        raise RuntimeError("Excel 为空或缺少表头行。")

    headers = [xnorm(h) for h in header_row]
    colmap = build_column_map(headers)

    written = 0
    row_idx = 0
    for row in it:
        if row is None:
            continue
        if all(xnorm(x) == "" for x in row):
            continue

        row_idx += 1
        meta = row_to_meta(row, colmap)

        filename = make_submission_filename(meta, row_idx)
        out_path = RAW_DIR / filename
        write_case_raw_md(meta, out_path)
        written += 1

    print(f"[xlsx->md] {xlsx_path.name} -> {written} cases_raw written into {RAW_DIR.as_posix()}")
    return written


# =========================================================
# 2) 原有 buildcases 逻辑（md -> site pages）
# =========================================================

def read_front_matter(md_text: str) -> dict:
    m = FRONT_MATTER_RE.match(md_text)
    if not m:
        return {}
    data = yaml.safe_load(m.group(1)) or {}
    return data if isinstance(data, dict) else {}


def strip_front_matter(md_text: str) -> str:
    m = FRONT_MATTER_RE.match(md_text)
    if not m:
        return md_text
    return md_text[m.end():]


def norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def display(v) -> str:
    """普通字段：空 -> NULL"""
    s = norm(v)
    return s if s else "NULL"


def display_nickname(v) -> str:
    """昵称字段：空 -> Anonymous"""
    s = norm(v)
    return s if s else "Anonymous"


def title_of(meta: dict) -> str:
    """
    页面标题规则：
    昵称 | 高考分数 | 录取院校 | 录取专业
    """
    return (
        f"{display_nickname(meta.get('nickname'))}"
        f" | {display(meta.get('gaokao_score'))}"
        f" | {display(meta.get('university'))}"
        f" | {display(meta.get('major'))}"
    )


def stable_slug(meta: dict, raw_path: Path) -> str:
    base = "|".join(
        [
            norm(meta.get("exam_year")),
            norm(meta.get("track")),
            norm(meta.get("nickname")),
            norm(meta.get("gaokao_score")),
            norm(meta.get("gaokao_rank")),
            norm(meta.get("university")),
            norm(meta.get("major")),
            raw_path.name,
        ]
    )
    h = hashlib.sha1(base.encode("utf-8")).hexdigest()[:10]
    return f"case-{h}"


def validate_track(meta: dict) -> None:
    t = norm(meta.get("track"))
    if t and t not in ("物理", "历史"):
        raise ValueError(f"track 必须为 '物理' 或 '历史'，当前为：{t}")


def case_link(stem: str) -> str:
    """
    关键修复：
    - 不要使用以 / 开头的绝对路径（GitHub Pages 项目页会丢掉 /<repo>/）
    - 使用相对链接：从 cases/index.md 指向同目录下的案例页面
    """
    return f"{stem}/"


def render_case_page(meta: dict, raw_path: Path, out_stem: str) -> str:
    """
    详情页展示规则：
    - 不出现“选填信息”小节标题
    - 保留（选填）注释
    - 选填为空展示 NULL
    - 昵称为空展示 Anonymous
    """
    t = title_of(meta)

    lines: list[str] = []
    lines.append("---")
    lines.append(f'title: "{t}"')
    lines.append("---")
    lines.append("")
    lines.append(f"# {t}")
    lines.append("")
    lines.append("## 基本信息")
    lines.append("")
    lines.append(f"- 昵称：{display_nickname(meta.get('nickname'))}")
    lines.append(f"- 考试年份：{display(meta.get('exam_year'))}")
    lines.append(f"- 选科：{display(meta.get('track'))}")
    lines.append(f"- 深一模校排（选填）：{display(meta.get('sz_mock1_rank'))}")
    lines.append(f"- 深二模校排（选填）：{display(meta.get('sz_mock2_rank'))}")
    lines.append(f"- 高考分数：{display(meta.get('gaokao_score'))}")
    lines.append(f"- 高考排名：{display(meta.get('gaokao_rank'))}")
    lines.append(f"- 录取院校：{display(meta.get('university'))}")
    lines.append(f"- 录取专业：{display(meta.get('major'))}")
    lines.append("")
    lines.append("## 院校评价（选填）")
    lines.append("")
    lines.append(display(meta.get("university_review")))
    lines.append("")
    lines.append("## 专业评价（选填）")
    lines.append("")
    lines.append(display(meta.get("major_review")))
    lines.append("")
    lines.append("## 给学弟学妹的建议（选填）")
    lines.append("")
    lines.append(display(meta.get("advice")))
    lines.append("")
    lines.append("> 备注：本案例由校友投稿整理，仅供参考。")
    lines.append(f"> 来源文件：`{raw_path.as_posix()}`")
    lines.append("")
    return "\n".join(lines)


def show_or_skip_null(text: str) -> str | None:
    """
    聚合页展示逻辑：
    - 单条评价为空：不展示该条（返回 None）
    - 若某个院校/专业下所有评价都为空：该院校/专业下显示 - NULL
    """
    t = text.strip() if isinstance(text, str) else str(text).strip()
    return t if t else None


def update_homepage_last_updated() -> None:
    """
    每次 build_cases 后更新 docs/index.md 中的“最后更新时间”一行：
    最后更新时间：YYYY/MM/DD  HH:MM:SS

    约定：docs/index.md 必须存在，且包含标记区块：
    <!-- LAST_UPDATED_START -->
    ...
    <!-- LAST_UPDATED_END -->
    """
    index_path = DOCS_DIR / "index.md"
    content = index_path.read_text(encoding="utf-8")

    start = "<!-- LAST_UPDATED_START -->"
    end = "<!-- LAST_UPDATED_END -->"
    if start not in content or end not in content:
        raise RuntimeError(
            "docs/index.md 缺少更新时间标记区块：\n"
            "<!-- LAST_UPDATED_START -->\n...\n<!-- LAST_UPDATED_END -->"
        )

    ts = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y/%m/%d  %H:%M:%S")
    line = f"最后更新时间：{ts}"
    new_block = f"{start}\n{line}\n{end}"

    content = re.sub(
        rf"{re.escape(start)}.*?{re.escape(end)}",
        new_block,
        content,
        flags=re.DOTALL,
    )
    index_path.write_text(content, encoding="utf-8")


def write_experience_page(cases: list[dict]) -> None:
    """
    生成 docs/experience.md
    汇总所有案例中的 advice（给学弟学妹的建议）。

    输出格式：
    Alan | 576 | 北京航空航天大学 | 通信工程：xxx

    规则：
    - 单条建议为空：不输出该条
    - 若全部为空：显示一行 NULL
    """
    lines: list[str] = []
    lines.append("# 查看经验")
    lines.append("")
    lines.append("本页汇总所有已收录案例的 **给学弟学妹的建议**。")
    lines.append("")

    shown = []
    for c in sorted(cases, key=lambda x: x["title"]):
        adv = show_or_skip_null(c.get("advice", ""))
        if adv is None:
            continue
        shown.append((c["title"], adv))

    if not shown:
        lines.append("- NULL")
    else:
        for title, adv in shown:
            lines.append(f"- **{title}**：{adv}")

    lines.append("")
    EXPERIENCE_FILE.write_text("\n".join(lines), encoding="utf-8")


def seniors_doc_title(md_path: Path) -> str:
    """
    取长文标题：
    1) front matter 的 title
    2) 第一个 H1（# xxx）
    3) 文件名（不含扩展名）
    """
    txt = md_path.read_text(encoding="utf-8")
    fm = read_front_matter(txt)
    t = norm(fm.get("title"))
    if t:
        return t

    body = strip_front_matter(txt)
    m = FIRST_H1_RE.search(body)
    if m:
        return m.group(1).strip()

    return md_path.stem


def write_seniors_index() -> None:
    """
    自动生成 docs/seniors/index.md
    - 扫描 docs/seniors/ 下所有 .md（排除 index.md）
    - 输出为可点击列表
    约定：docs/seniors/ 目录必须存在（若不存在将直接报错，符合你的要求）
    """
    if not SENIORS_DIR.exists():
        raise RuntimeError("缺少目录：docs/seniors/（请先创建该目录）")

    posts = []
    for p in sorted(SENIORS_DIR.glob("*.md")):
        if p.name.lower() == "index.md":
            continue
        title = seniors_doc_title(p)
        # 关键修复：使用相对链接（从 seniors/index.md 指向同目录文章）
        link = f"{p.stem}/"
        posts.append((title, link))

    lines: list[str] = []
    lines.append("# 学长学姐说")
    lines.append("")
    lines.append("本栏目收录来自学长学姐投稿的**长文分享**。")
    lines.append("")

    if not posts:
        lines.append("当前暂无长文投稿。你可以将长文 Markdown 放入 `docs/seniors/` 目录后重新构建。")
        lines.append("")
        SENIORS_INDEX_FILE.write_text("\n".join(lines), encoding="utf-8")
        return

    lines.append(f"当前收录：**{len(posts)}** 篇。文章以投稿时间由新到旧排序。")
    lines.append("")
    for title, link in sorted(posts, key=lambda x: x[0]):
        lines.append(f"- [{title}]({link})")
    lines.append("")
    SENIORS_INDEX_FILE.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    # 先导入 xlsx -> cases_raw
    import_xlsx_to_cases_raw()

    # 继续原构建
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    raw_files = sorted(RAW_DIR.glob("*.md"))
    cases: list[dict] = []

    for raw in raw_files:
        text = raw.read_text(encoding="utf-8")
        meta = read_front_matter(text) or {}

        try:
            validate_track(meta)
        except ValueError as e:
            meta_err = dict(meta)
            out_stem = stable_slug(meta_err, raw)
            out_path = OUT_DIR / f"{out_stem}.md"
            out_path.write_text(
                f'---\ntitle: "{title_of(meta_err)}"\n---\n\n# {title_of(meta_err)}\n\n**字段校验失败：** {e}\n',
                encoding="utf-8",
            )
            cases.append(
                {
                    "title": title_of(meta_err),
                    "stem": out_stem,
                    "nickname": display_nickname(meta_err.get("nickname")),
                    "university": display(meta_err.get("university")),
                    "major": display(meta_err.get("major")),
                    "university_review": norm(meta_err.get("university_review")),
                    "major_review": norm(meta_err.get("major_review")),
                    "advice": norm(meta_err.get("advice")),
                }
            )
            continue

        out_stem = stable_slug(meta, raw)
        out_path = OUT_DIR / f"{out_stem}.md"
        out_path.write_text(render_case_page(meta, raw, out_stem), encoding="utf-8")

        cases.append(
            {
                "title": title_of(meta),
                "stem": out_stem,
                "nickname": display_nickname(meta.get("nickname")),
                "university": display(meta.get("university")),
                "major": display(meta.get("major")),
                "university_review": norm(meta.get("university_review")),
                "major_review": norm(meta.get("major_review")),
                "advice": norm(meta.get("advice")),
            }
        )

    # ---------- 案例总览 ----------
    cases_sorted = sorted(cases, key=lambda c: c["title"])
    lines = []
    lines.append("# 案例总览")
    lines.append("")
    lines.append(f"当前收录：**{len(cases_sorted)}** 条。点击标题进入详情页。")
    lines.append("")
    for c in cases_sorted:
        lines.append(f"- [{c['title']}]({case_link(c['stem'])})")
    lines.append("")
    INDEX_FILE.write_text("\n".join(lines), encoding="utf-8")

    # ---------- 按院校 ----------
    uni_map: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for c in cases:
        uni = c["university"]
        nick = c["nickname"]
        maj = c["major"]
        review = c.get("university_review", "")
        uni_map[uni].append((f"{nick} | {maj}", review))

    lines = []
    lines.append("# 按院校")
    lines.append("")
    lines.append("展示该院校的**院校评价**聚合结果，格式为 `昵称 | 专业：评价`。")
    lines.append("")
    for uni in sorted(uni_map.keys()):
        items = uni_map[uni]
        shown: list[tuple[str, str]] = []
        for prefix, review in items:
            txt = show_or_skip_null(review)
            if txt is not None:
                shown.append((prefix, txt))

        lines.append(f"## {uni}（{len(shown)}）")
        lines.append("")
        if not shown:
            lines.append("- NULL")
        else:
            for prefix, txt in shown:
                lines.append(f"- **{prefix}**：{txt}")
        lines.append("")
    BY_UNI_FILE.write_text("\n".join(lines), encoding="utf-8")

    # ---------- 按专业 ----------
    major_map: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for c in cases:
        maj = c["major"]
        nick = c["nickname"]
        uni = c["university"]
        review = c.get("major_review", "")
        major_map[maj].append((f"{nick} | {uni}", review))

    lines = []
    lines.append("# 按专业")
    lines.append("")
    lines.append("展示该专业的**专业评价**聚合结果，格式为 `昵称 | 院校：评价`。")
    lines.append("")
    for maj in sorted(major_map.keys()):
        items = major_map[maj]
        shown: list[tuple[str, str]] = []
        for prefix, review in items:
            txt = show_or_skip_null(review)
            if txt is not None:
                shown.append((prefix, txt))

        lines.append(f"## {maj}（{len(shown)}）")
        lines.append("")
        if not shown:
            lines.append("- NULL")
        else:
            for prefix, txt in shown:
                lines.append(f"- **{prefix}**：{txt}")
        lines.append("")
    BY_MAJOR_FILE.write_text("\n".join(lines), encoding="utf-8")

    # ---------- 查看经验（汇总 advice） ----------
    write_experience_page(cases)

    # ---------- 学长学姐说：自动生成 seniors 目录索引 ----------
    write_seniors_index()

    # ---------- 更新首页“最后更新时间” ----------
    update_homepage_last_updated()

    print(
        f"Built {len(cases)} cases. "
        f"Generated: cases/index.md, cases/by-university.md, cases/by-major.md, experience.md, seniors/index.md"
    )


if __name__ == "__main__":
    main()